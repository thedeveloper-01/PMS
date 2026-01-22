"""
Excel export functionality using openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from payroll_system.models.attendance import Attendance
from payroll_system.models.payroll import Payroll
from payroll_system.models.employee import Employee
from payroll_system.repository.employee_repository import EmployeeRepository
from payroll_system.config import EXPORTS_DIR
from datetime import datetime
import os

class ExcelExporter:
    """Export payroll data to Excel"""
    
    def __init__(self):
        self.employee_repo = EmployeeRepository()
    
    def export_payroll_report(self, payrolls: List[Payroll], 
                            month: int, year: int) -> str:
        """Export payroll report to Excel"""
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Payroll_{year}_{month:02d}"
            
            # Define styles
            header_fill = PatternFill(start_color="5C7EB5", end_color="5C7EB5", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                'Employee ID', 'Employee Name', 'Basic Salary', 'HRA', 'DA',
                'Allowances', 'Bonus', 'Overtime Pay', 'Gross Salary',
                'PF', 'ESI', 'PT', 'LOP Deduction', 'Total Deductions', 'Net Salary',
                'Present Days', 'Working Days', 'LOP Days'
            ]
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Write data
            for row_num, payroll in enumerate(payrolls, 2):
                employee = self.employee_repo.get_by_id(payroll.employee_id)
                employee_name = employee.employee_name if employee else "N/A"
                
                row_data = [
                    payroll.employee_id,
                    employee_name,
                    payroll.basic_salary,
                    payroll.hra,
                    payroll.da,
                    payroll.allowances,
                    payroll.bonus,
                    payroll.overtime_pay,
                    payroll.gross_salary,
                    payroll.pf,
                    payroll.esi,
                    payroll.pt,
                    payroll.lop_deduction,
                    payroll.total_deductions,
                    payroll.net_salary,
                    payroll.present_days,
                    payroll.working_days,
                    payroll.lop_days
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if col_num >= 3:  # Numeric columns
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-adjust column widths
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                for row in ws[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary row
            summary_row = len(payrolls) + 3
            ws.cell(row=summary_row, column=1).value = "TOTAL"
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            
            # Calculate totals
            total_cols = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
            for col_num in total_cols:
                col_letter = get_column_letter(col_num)
                formula = f"=SUM({col_letter}2:{col_letter}{len(payrolls) + 1})"
                cell = ws.cell(row=summary_row, column=col_num)
                cell.value = formula
                cell.font = Font(bold=True)
                cell.number_format = '#,##0.00'
                cell.border = border
            
            # Save file
            filename = f"payroll_report_{year}_{month:02d}.xlsx"
            filepath = EXPORTS_DIR / filename
            wb.save(str(filepath))
            
            return str(filepath)
        except Exception as e:
            raise Exception(f"Error exporting to Excel: {str(e)}")
    
    def export_employee_list(self, employees: List[Employee]) -> str:
        """Export employee list to Excel"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Employees"
            
            # Styles
            header_fill = PatternFill(start_color="5C7EB5", end_color="5C7EB5", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                'Employee ID', 'Employee Name', 'Email', 'Mobile', 'Gender',
                'Department', 'Designation', 'Branch', 'Basic Salary', 'Status'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Write data
            for row_num, employee in enumerate(employees, 2):
                row_data = [
                    employee.employee_id,
                    employee.employee_name,
                    employee.email,
                    employee.mobile_number,
                    employee.gender,
                    employee.department_id or "N/A",
                    employee.designation_id or "N/A",
                    employee.branch_id or "N/A",
                    employee.basic_salary,
                    "Active" if employee.status == 1 else "Inactive"
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if col_num == 9:  # Salary column
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-adjust column widths
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                for row in ws[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            filename = f"employee_list_{datetime.now().strftime('%Y%m%d')}.xlsx"
            filepath = EXPORTS_DIR / filename
            wb.save(str(filepath))
            
            return str(filepath)
        except Exception as e:
            raise Exception(f"Error exporting employee list: {str(e)}")

    def export_attendance_report(self, attendances: List[Attendance], month: int, year: int) -> str:
        """Export attendance report to Excel"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Attendance_{year}_{month:02d}"
            
            # Styles
            header_fill = PatternFill(start_color="5C7EB5", end_color="5C7EB5", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Headers
            headers = ['Date', 'Employee ID', 'Name', 'Check In', 'Check Out', 'Status']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Write data
            for row_num, att in enumerate(attendances, 2):
                employee = self.employee_repo.get_by_id(att.employee_id)
                name = employee.employee_name if employee else "N/A"
                
                row_data = [
                    att.date,
                    att.employee_id,
                    name,
                    att.checkin_time,
                    att.checkout_time,
                    att.status
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Auto-adjust column widths
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                for row in ws[column_letter]:
                    try:
                        if len(str(row.value)) > max_length:
                            max_length = len(str(row.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f"attendance_report_{year}_{month:02d}.xlsx"
            filepath = EXPORTS_DIR / filename
            wb.save(str(filepath))
            
            return str(filepath)
        except Exception as e:
            raise Exception(f"Error exporting attendance: {str(e)}")

    def generate_salary_summary(self, payrolls: List[Payroll], year: int) -> str:
        """Generate annual salary summary"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Salary_Summary_{year}"
            
            # Styles
            header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Headers
            headers = ['Employee ID', 'Name', 'Total Gross', 'Total Deductions', 'Total Net Salary', 'Total Bonus']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Aggregate data by employee
            summary = {}
            for p in payrolls:
                if p.employee_id not in summary:
                    employee = self.employee_repo.get_by_id(p.employee_id)
                    summary[p.employee_id] = {
                        'name': employee.employee_name if employee else "N/A",
                        'gross': 0, 'deductions': 0, 'net': 0, 'bonus': 0
                    }
                summary[p.employee_id]['gross'] += p.gross_salary
                summary[p.employee_id]['deductions'] += p.total_deductions
                summary[p.employee_id]['net'] += p.net_salary
                summary[p.employee_id]['bonus'] += p.bonus
            
            # Write data
            row_num = 2
            for emp_id, data in summary.items():
                row_data = [
                    emp_id, data['name'], data['gross'], data['deductions'], data['net'], data['bonus']
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if col_num > 2:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                row_num += 1
            
            # Auto-adjust column widths
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = 25
            
            filename = f"salary_summary_{year}.xlsx"
            filepath = EXPORTS_DIR / filename
            wb.save(str(filepath))
            
            return str(filepath)
        except Exception as e:
            raise Exception(f"Error generating summary: {str(e)}")

